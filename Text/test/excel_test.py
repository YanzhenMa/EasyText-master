# encoding=utf-8
import time

import numpy as np
import openpyxl
import pandas as pd
from Text.test.all_features import Features


class Analysis:

    def __init__(self, path):
        self.file_path = path

    def analyze_excel(self):
        '''分析诗歌excels
        '''
        data = pd.read_excel(self.file_path, sheet_name=0)
        data = data.replace(np.nan, '')
        length = len(data.index.values)
        # 写入excel表头
        names = ['字数', '字种', '字频1', '字频2', '笔画1', '笔画2', '常用字比例2500', '常用字种2500', '常用字比例3500', '常用字种3500', '通用字比例7000', '通用字种7000', '平均构词能力', '词数', '词种数', '词频1', '词频2', '平均词长', '难词数', '难词比例', '难词词种比例', '名词数', '动词数', '形容词数', '数词数', '量词数', '代词数', '实词数', '介词数', '连词数', '助词数', '叹词数', '语气词数', '时序词数', '虚词数', '实词词种数',
                 '虚词词种数', '实词比例', '实词词种比例', '虚词比例', '虚词词种比例', '平均名词修饰语数量', '平均谓语修饰语数量', '一级词数量', '二级词数量', '三级词数量', '四级词数量', '五级词数量', '句子数', '平均句子字数含标点', '平均句子字数不含标点', '最短句子字数', '最长句子字数', '百字句子所占比例', '相邻句子名词重叠平均数', '相邻句子名词重叠标准差', '所有句子名词重叠平均数', '所有句子名词重叠标准差', '相邻句子实词重叠平均数', '相邻句子实词重叠标准差', '所有句子实词重叠平均数', '所有句子实词重叠标准差']
        mywb = openpyxl.load_workbook(self.file_path)
        sheet = mywb['Sheet1']
        col = 6
        row = 1
        for i in range(len(names)):
            sheet.cell(row, col+i, str(names[i]))
        mywb.save(self.file_path)
        mywb.close()
        # 写入计算结果
        start_index = 0
        for i in range(start_index, length):
            print(i)
            t = time.localtime()
            print(time.strftime("%Y-%m-%d %H:%M:%S", t))
            astr = data['原文'][i]
            mywb = openpyxl.load_workbook(self.file_path)
            sheet = mywb['Sheet1']
            row = i+2
            if astr:
                feat = Features(astr)
                data = feat.get_features()
                values = list(data.values())
                for j in range(0, len(values)):
                    sheet.cell(row, col+j, str(values[j]))
            else:
                pass
            mywb.save(self.file_path)
            mywb.close()


if __name__ == '__main__':
    list_names = ['北师大版', '人教版', '苏教版', '统编版']
    for name in list_names:
        file_path = r'E:\eachina\easytext\Text\test\test_data\{}补全诗词.xlsx'.format(
            name)
        print(file_path)
        obj = Analysis(file_path)
        obj.analyze_excel()
